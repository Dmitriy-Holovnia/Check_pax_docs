import openpyxl
import datetime
from iso3166 import countries

wb = openpyxl.load_workbook('./test.xlsx')
data_sheet = wb["Entry data"]
final_sheet = wb["Final data"]

total_pax = 179

# <<=== CODE ===>>
# Remake booking date
def go_booking():
    for row in range(2, total_pax + 1):
        date = data_sheet.cell(row=row, column=1).value
        txt = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S').date().strftime("%d/%m/%y")
        final_sheet.cell(row=row, column=1, value=txt)

# Remake given name
def go_given_name():
    for row in range(2, total_pax + 1):
        txt = data_sheet.cell(row=row, column=2).value
        index = txt.find("/")
        name = txt[index + 1:]
        final_sheet.cell(row=row, column=2, value=name)

# Remake surname
def go_surname():
    for row in range(2, total_pax + 1):
        txt = data_sheet.cell(row=row, column=2).value
        index = txt.find("/")
        name = txt[:index]
        final_sheet.cell(row=row, column=3, value=name)

# Remake surname
def go_gender():
    for row in range(2, total_pax + 1):
        txt = data_sheet.cell(row=row, column=3).value
        if txt == "male":
            gender = "M"
        else:
            gender = "F"
        final_sheet.cell(row=row, column=4, value=gender)

# Remake nationality
def go_nationality():
    for row in range(2, total_pax + 1):
        nationality = data_sheet.cell(row=row, column=4).value
        try:
            code = countries.get(nationality).alpha3
            final_sheet.cell(row=row, column=5, value=code)
        except:
            print("""
            ==============================
            Error nationality for {} in row {}
            ===============================
            """.format(nationality, row))

# - Remake expirity date -
def go_expirity_date():
    for row in range(2, total_pax + 1):
        txt = data_sheet.cell(row=row, column=5).value
        date = datetime.datetime.strptime(txt, '%Y-%m-%d').date().strftime("%d/%m/%Y")
        final_sheet.cell(row=row, column=6, value=date)

# - Remake birthday -
def go_birthday():
    for row in range(2, total_pax + 1):
        txt = data_sheet.cell(row=row, column=6).value
        date = datetime.datetime.strptime(txt, '%Y-%m-%d').date().strftime("%d/%m/%Y")
        final_sheet.cell(row=row, column=7, value=date)

# - Main function -
def remake_data():
    for column in range(1, 8):
        if column == 1:
            go_booking()
        elif column == 2:
            go_given_name()
        elif column == 3:
            go_surname()
        elif column == 4:
            go_gender()
        elif column == 5:
            go_nationality()
        elif column == 6:
            go_expirity_date()
        elif column == 7:
            go_birthday()
    wb.save(filename="test.xlsx")
    print("Data remake was successful")
remake_data()



